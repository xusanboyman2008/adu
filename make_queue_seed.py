#!/usr/bin/env python3
"""Build the coordinator queue seed from names.xlsx, pre-marking people who are
already done so the PCs never redo them.

It outputs queue_seed.csv with the FULL coordinator column layout (so you import
it once and you're ready). Students that already completed — matched by full name
against your Google Form responses sheet — are written as status=done together
with their existing email/password/certificate URL. Everyone else is pending.

Usage:
    # roster only (everyone pending):
    python3 make_queue_seed.py

    # roster + already-done people from the responses Google Sheet (or a CSV):
    python3 make_queue_seed.py --done "https://docs.google.com/spreadsheets/d/<ID>/edit"
    python3 make_queue_seed.py --done existing.csv

    # options:
    python3 make_queue_seed.py --names names.xlsx --out queue_seed.csv --done <url|csv>

The --done source must have columns: FULL NAME, Email, Password, URL
(the standard Google Form response layout).
"""

import argparse
import csv
import io
import re
import sys
import urllib.request

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")

SHEET = "Talabalar"
# Coordinator column order — must match coordinator/Code.gs.
HEADER = [
    "student_id", "full_name", "status", "owner", "attempts", "claimed_at",
    "lease_expires", "finished_at", "last_error", "email", "password", "certificate_url",
]


def norm(s):
    """Match key: uppercase, strip everything but letters/digits (so spacing and
    apostrophe variants like O' / O‘ / O` don't matter)."""
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def load_done(source):
    """Return {normalized_full_name: (email, password, cert_url)} from a Google
    Sheets URL/ID or a local CSV with FULL NAME / Email / Password / URL columns."""
    text = _read_done_source(source)
    done = {}
    reader = csv.DictReader(io.StringIO(text))
    # tolerate header case/spacing differences
    lut = {re.sub(r"[^a-z]", "", (h or "").lower()): h for h in (reader.fieldnames or [])}
    name_col = lut.get("fullname")
    email_col = lut.get("email")
    pass_col = lut.get("password")
    url_col = lut.get("url") or lut.get("certificateurl")
    if not name_col:
        sys.exit(f"--done source has no 'FULL NAME' column. Found: {reader.fieldnames}")
    for row in reader:
        nm = row.get(name_col, "")
        if not str(nm).strip():
            continue
        done[norm(nm)] = (
            (row.get(email_col) or "").strip() if email_col else "",
            (row.get(pass_col) or "").strip() if pass_col else "",
            (row.get(url_col) or "").strip() if url_col else "",
        )
    return done


def _read_done_source(source):
    # A Google Sheets link or bare ID -> download the CSV export of its first tab.
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", source)
    sheet_id = m.group(1) if m else (source if re.fullmatch(r"[a-zA-Z0-9-_]{30,}", source) else None)
    if sheet_id:
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        print(f"Downloading already-done responses from sheet {sheet_id} ...")
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req) as resp:
            return resp.read().decode("utf-8", "replace")
    # Otherwise treat as a local CSV file path.
    with open(source, encoding="utf-8") as fh:
        return fh.read()


def load_roster(names_file):
    wb = openpyxl.load_workbook(names_file, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET}' not found in {names_file}. Tabs: {wb.sheetnames}")
    sheet = wb[SHEET]
    roster = []
    for r in range(3, sheet.max_row + 1):
        sid, name = sheet.cell(r, 1).value, sheet.cell(r, 2).value
        if sid and name:
            roster.append((str(sid).strip(), str(name).strip()))
    return roster


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--names", default="names.xlsx")
    ap.add_argument("--out", default="queue_seed.csv")
    ap.add_argument("--done", default=None,
                    help="Google Sheets URL/ID or CSV of already-completed people")
    ap.add_argument("--skip-first", type=int, default=0,
                    help="Reserve the first N students (by roster order) so the "
                         "queue never hands them out — leave them to your old "
                         "START/END script. They are marked 'reserved'.")
    # positional back-compat: make_queue_seed.py [names.xlsx] [out.csv]
    ap.add_argument("pos", nargs="*")
    args = ap.parse_args()
    names_file = args.pos[0] if len(args.pos) > 0 else args.names
    out_file = args.pos[1] if len(args.pos) > 1 else args.out

    roster = load_roster(names_file)
    done = load_done(args.done) if args.done else {}
    skip_first = max(0, args.skip_first)

    n_done = 0
    n_reserved = 0
    rows = []
    for idx, (sid, name) in enumerate(roster, start=1):  # 1-based, matches START/END
        d = done.get(norm(name))
        if d:
            # Already completed: keep it 'done' (with creds) wherever it is.
            email, password, cert = d
            rows.append([sid, name, "done", "", 1, "", "", "", "", email, password, cert])
            n_done += 1
        elif idx <= skip_first:
            # First N students belong to the old script — never claim them.
            rows.append([sid, name, "reserved", "", 0, "", "", "", "", "", "", ""])
            n_reserved += 1
        else:
            rows.append([sid, name, "pending", "", 0, "", "", "", "", "", "", ""])

    with open(out_file, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(HEADER)
        w.writerows(rows)

    print(f"Wrote {len(rows)} students to {out_file}")
    print(f"  already done (skipped)      : {n_done}")
    if n_reserved:
        print(f"  reserved for old script     : {n_reserved} (first {skip_first} by order)")
    print(f"  to process by the queue     : {len(rows) - n_done - n_reserved}")
    if done:
        matched_names = sum(1 for _, name in roster if norm(name) in done)
        unmatched = len(done) - len({norm(name) for _, name in roster if norm(name) in done})
        if unmatched:
            print(f"  note: {unmatched} completed name(s) in --done weren't found in the roster")
    print("\nNext: in your coordinator Google Sheet, click cell A1, then "
          "File > Import > Upload this CSV > 'Replace data at selected cell'. "
          "The status column is already filled, so you can skip initQueue "
          "(or run it — it only fills blank statuses).")


if __name__ == "__main__":
    main()
